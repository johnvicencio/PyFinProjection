import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles import Font, Alignment

# Initialize the DataFrame with predefined columns
columns = ["Category", "Item"] + [f"{month}" for month in ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]] + ["Total", "Rate"]
df = pd.DataFrame(columns=columns)

# Define company name and subtitle
company_name = "Your Company Name"
subtitle = "Year 1 Profit and Loss Projection Statement"

# Define projection percentage and projected revenue
projection_percent = 1  # 1.05 is for Projecting a 5% increase
projected_revenue = 500000  # $500K

# Define fixed expense categories
fixed_expense_items = [
    "Rent", "JB Mgt", "Insurance - Medical", "Insurance - Business", "Management",
    "Note Payment", "Taxes", "Employee Morale", "Equipment Lease", "Repair & Maintenance",
    "Depreciation", "Total Fixed Expense"
]

# Function to create a row with seasonal rate adjustment for variable expenses
def create_row(category, item, rate, seasonal_rates=None):
    if seasonal_rates is None:
        seasonal_rates = {month: 1.0 for month in ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]}
    
    if item in fixed_expense_items:
        # Fixed expenses are not affected by seasonal rates
        monthly_value = (projected_revenue * rate) / 12
        total = monthly_value * 12
        row = {"Category": category, "Item": item, "Rate": rate}  # No multiplication by 100 here
        row.update({month: monthly_value for month in ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]})
        row["Total"] = total
    else:
        # Variable expenses with seasonal rates
        monthly_values = {}
        total = 0
        for month in ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]:
            monthly_value = (projected_revenue * rate * seasonal_rates[month]) / 12
            monthly_values[month] = monthly_value
            total += monthly_value
        
        row = {"Category": category, "Item": item, "Rate": rate}  # No multiplication by 100 here
        row.update(monthly_values)
        row["Total"] = total
    
    return row



# List to store DataFrames for each row
rows = []

# Function to add a blank row
def add_blank_row():
    return pd.DataFrame([[""] * len(columns)], columns=columns)

# Function to add rows for a category
def add_category_rows(category, items, seasonal_rates=None):
    rows.append(pd.DataFrame([{"Category": category, "Item": "", "Rate": "", **{month: "" for month in ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]}, "Total": ""}]))
    for item, rate in items:
        rows.append(pd.DataFrame([create_row("", item, rate, seasonal_rates)]))
    rows.append(add_blank_row())  # Blank row after the category totals

# Seasonal rates for each month
seasonal_rates = {
    "Jan": 0.8,
    "Feb": 0.9,
    "Mar": 1.0,
    "Apr": 1.1,
    "May": 1.2,
    "Jun": 1.3,
    "Jul": 1.4,
    "Aug": 1.3,
    "Sep": 1.2,
    "Oct": 1.1,
    "Nov": 1.0,
    "Dec": 0.9
}

# Add rows for each item under categories with seasonal rates
add_category_rows("REVENUE / SALES", [
    ("Gross Sales", projection_percent),
    ("Sales Return & Discounts", 0.0),
    ("Net Sales", projection_percent)
], seasonal_rates)

add_category_rows("COGS", [
    ("Food Cost", 0.275),
    ("Paper Cost", 0.02),
    ("Labor Cost", 0.245),
    ("Supplies Other", 0.003),
    ("Total COGS", 0.543)
], seasonal_rates)

add_category_rows("STORE EXPENSES", [
    ("Rent", 0.144),
    ("JB Mgt", 0.09),
    ("Insurance - Medical", 0.0),
    ("Insurance - Business", 0.005),
    ("Management", 0.048),
    ("Note Payment", 0.0),
    ("Taxes", 0.001),
    ("Employee Morale", 0.0),
    ("Equipment Lease", 0.0),
    ("Repair & Maintenance", 0.002),
    ("Depreciation", 0.289),
    ("Total Fixed Expense", 0.579)
], seasonal_rates)

add_category_rows("", [
    ("Payroll Tax", 0.011),
    ("Uniforms", 0.001),
    ("Salaries", 0.0),
    ("Employee Incentive", 0.0),
    ("Total Compensation", 0.012)
], seasonal_rates)

add_category_rows("", [
    ("Marketing/Advertising", 0.0),
    ("Training", 0.0),
    ("Utilities", 0.009),
    ("Postage", 0.0),
    ("Cash Over/Short", 0.0),
    ("Services Other", 0.0),
    ("Credit Card Fees", 0.015),
    ("Supplies", 0.005),
    ("Other", 0.003),
    ("Total Variable Expense", 0.032)
], seasonal_rates)

add_category_rows("", [
    ("Total Expenses", 0.333),
    ("Store Contribution", 0.155),
    ("Store EBITDA", 0.125)
], seasonal_rates)

add_category_rows("G&A", [
    ("Salaries & Wages (Owner)", 0.10),
    ("Payroll Taxes", 0.005),
    ("Total Compensation", 0.105)
], seasonal_rates)

add_category_rows("", [
    ("Temporary Labor", 0.0),
    ("Travel & Entertainment", 0.0),
    ("Computer", 0.0),
    ("Telephone", 0.0),
    ("Bank Fees", 0.0),
    ("Postage & Delivery", 0.0),
    ("Office Supplies", 0.0),
    ("Internet", 0.0),
    ("Accounting & Finance", 0.0),
    ("Storage", 0.0),
    ("Misc.", 0.0),
    ("Total G&A Expenses", 0.105)
], seasonal_rates)

add_category_rows("", [
    ("Total Corporate Overhead", 0.105),
    ("Total EBITDA", 0.105)
], seasonal_rates)

# Concatenate all rows into a single DataFrame
df = pd.concat(rows, ignore_index=True)

# Reorder columns to move 'Rate' to the end
df = df[[col for col in columns if col != "Rate"] + ["Rate"]]

# Save DataFrame to Excel
excel_path = "financial_statement.xlsx"
df.to_excel(excel_path, index=False)

# Open the saved Excel file to adjust column widths and add formatting
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Insert two new rows at the top
ws.insert_rows(1, 2)

# Merge cells for the company name and subtitle
ws.merge_cells('A1:P1')
ws.merge_cells('A2:P2')

# Set company name and subtitle with bold and centered alignment
ws['A1'] = company_name
ws['A1'].font = Font(bold=True, size=16)  
ws['A1'].alignment = Alignment(horizontal='center')

ws['A2'] = subtitle
ws['A2'].font = Font(bold=True, size=14)  
ws['A2'].alignment = Alignment(horizontal='center')

# Adjust column widths
column_widths = {
    "A": 14,
    "B": 22,
    "C": 13.5,
    "D": 13.5,
    "E": 13.5,
    "F": 13.5,
    "G": 13.5,
    "H": 13.5,
    "I": 13.5,
    "J": 13.5,
    "K": 13.5,
    "L": 13.5,
    "M": 13.5,
    "N": 13.5,
    "O": 13.5,
    "P": 10
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Format columns C to N for 2 decimal places
for col_index in range(2, 14):  # Columns C to N (Excel uses 1-based index)
    col_letter = openpyxl.utils.get_column_letter(col_index)
    for cell in ws[col_letter]:
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'

# Format the "Rate" column with percentage
for cell in ws['P']:
    if cell.value and isinstance(cell.value, (int, float)):
        cell.number_format = '0.0%'

# Bold the Category rows and any row with "Total" or "Net Sales"
for row in ws.iter_rows(min_row=1, max_col=len(columns), values_only=False):
    # Check if "Net Sales" is in any of the cells of the row
    if any(cell.value and "Net Sales" in str(cell.value) for cell in row):
        for cell in row:
            cell.font = Font(bold=True)
    elif row[0].value and (row[0].value.startswith("REVENUE / SALES") or \
                           row[0].value.startswith("COGS") or \
                           row[0].value.startswith("STORE EXPENSES") or \
                           row[0].value.startswith("G&A")):
        for cell in row:
            cell.font = Font(bold=True)
    elif any(cell.value and "Total" in str(cell.value) for cell in row):
        for cell in row:
            cell.font = Font(bold=True)


# Save the modified Excel file
wb.save(excel_path)

print(f"Financial statement saved to {excel_path}")
